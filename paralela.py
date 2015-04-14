import random
from openpyxl import load_workbook


def Voto(izquierda, derecha, independiente):
    rango1 = izquierda
    rango2 = izquierda + derecha
    probabilidad = random.randint(0, 100)
    voto = 0
    if probabilidad <= rango1:
        voto = 1
    elif probabilidad > rango1 and probabilidad <= rango2:
        voto = 2
    elif probabilidad > rango2 and probabilidad <= 100:
        voto = 3
    return voto


def LeerComunas():
    wb = load_workbook('comunas.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')
    lista = []
    for i in range(2, 349):
        comuna = ws.cell('A' + str(i))
        region = ws.cell('B' + str(i))
        derecha = ws.cell('C' + str(i))
        izquierda = ws.cell('D' + str(i))
        independiente = ws.cell('E' + str(i))
        dato = [comuna.value, region.value, derecha.value,
                izquierda.value, independiente.value]
        lista.append(dato)
    return lista


def contarVoto(comunas):
    wb = load_workbook('data_servel.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')
    suma = 0
    sumb = 0
    sumc = 0
    for i in range(2, 200002):
        rut = ws.cell('A' + str(i))
        rut = str(rut.value)
        comuna = ws.cell('B' + str(i))
        for j in range(len(comunas)):
            if str(comuna.value) == comunas[j][0]:
                voto = Voto(comunas[j][2], comunas[j][3], comunas[j][4])
                if voto == 1:
                    suma = suma + 1
                elif voto == 2:
                    sumb = sumb + 1
                elif voto == 3:
                    sumc = sumc + 1
    print "izquierda: " + str(suma)
    print "derecha: " + str(sumb)
    print "independiente: " + str(sumc)

comunas = LeerComunas()
contarVoto(comunas)
